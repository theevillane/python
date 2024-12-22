import hashlib
import re
from openpyxl import Workbook, load_workbook
from getpass import getpass
import os
import logging
from datetime import datetime
import random
from functools import wraps
import smtplib
from email.mime.text import MIMEText


logging.basicConfig(
    filename='app_debug.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

EXCEL_FILE = "attendance_db.xlsx"
DEFAULT_PASSWORD = "password123"
ADMIN_CONFIRM_PASSWORD = hashlib.sha256("Kijanamdogo".encode()).hexdigest()

attendance_list = []

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        
        # Users Sheet
        users_sheet = wb.active
        users_sheet.title = "Users"
        users_sheet.append(["Username", "Password", "Role", "Email"])

        # Attendance Sheet
        attendance_sheet = wb.create_sheet("Attendance")
        attendance_sheet.append(["Username", "Date", "Check-in-Time", "Check-out-Time", "Total Hours"])

        # Audit Log Sheet
        audit_sheet = wb.create_sheet("Audit Log")
        audit_sheet.append(["Timestamp", "Username", "Action", "Details"])

        wb.save(EXCEL_FILE)
        print("Excel db initialized.")

def hash_password(password):
    """Hash a password for secure storage."""
    return hashlib.sha256(password.encode()).hexdigest()

def is_strong_password(password):
    """Check if the password is strong."""
    return (len(password) >= 8 and
            re.search(r"[A-Za-z]", password) and
            re.search(r"[0-9]", password) and
            re.search(r"[!@#$%^&*(),.?\":{}|<>]", password))

def load_sheet(sheet_name):
    """Load a specific sheet from the excel file."""
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb[sheet_name]
        return sheet, wb
    except KeyError:
        print(f"Sheet '{sheet_name}' not found")
    except FileNotFoundError:
        print(f"File '{EXCEL_FILE}' not found.")
    return None, None

def save_to_excel(workbook):
    """Save changes to Excel file."""
    try:
        workbook.save(EXCEL_FILE)
    except Exception as e:
        logging.error(f"Error saving workbook: {e}")

def send_email_notification(to_email, subject, message):
    """Send an email notification with timestamp."""
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        message += f"\n\nAction Time: {timestamp}"

        from_email = os.getenv('EMAIL_ADDRESS')
        from_password = os.getenv('EMAIL_PASSWORD')
        if not from_email or not from_password:
           print("Email credentials are missing. Notifications will not be sent.")
           return

        smtp_server = "smtp.gmail.com"
        smtp_port = 587

        msg = MIMEText(message)
        msg['Subject'] = subject
        msg['From'] = from_email
        msg['To'] = to_email

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(from_email, from_password)
            server.sendmail(from_email, to_email, msg.as_string())

        print(f"Notification sent to {to_email} at {timestamp}.")
    except Exception as e:
        print(f"Failed to send email: {e}")

def add_check_in(username):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet, wb = load_sheet("Attendance")
    if not sheet:
        print("Error loading attendance data.")
        return

    sheet.append([username, datetime.now().date(), timestamp, "", ""])
    save_to_excel(wb)
    print(f"Check-in recorded for {username} at {timestamp}.")

def add_check_out(username):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet, wb = load_sheet("Attendance")
    if not sheet:
        print("Error loading attendance data.")
        return

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == username and not row[3].value:
            row[3].value = timestamp
            check_in = datetime.strptime(row[2].value, '%Y-%m-%d %H:%M:%S')
            check_out = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            row[4].value = str(check_out - check_in)
            save_to_excel(wb)
            print(f"Check-out recorded for {username} at {timestamp}.")
            return
    print("No check-in record found or already checked out.")

def log_audit_action(username, action, details=""):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet, wb = load_sheet("Audit Log")
    if not sheet:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.create_sheet("Audit Log")
        sheet.append(["Timestamp", "Username", "Action", "Details"])

    sheet.append([timestamp, username, action, details])
    save_to_excel(wb)
    logging.info(f"{username} | {action} | {details} | {timestamp}")


def generate_otp():
    """Generate a 6-digit OTP."""
    return random.randint(100000, 999999)

def verify_otp(user_otp, generated_otp):
    """Verify the OTP entered by the user."""
    return user_otp == generated_otp

def require_role(required_role):
    """Decorator to enforce role-based access."""
    def decorator(func):
        @wraps(func)
        def wrapper(username, role, *args, **kwargs):
            if role != required_role:
                print(f"Access denied. You must be a {required_role} to perform this action.")
                return
            return func(username, role, *args, **kwargs)
        return wrapper
    return decorator

def register_user():
    """Register a new user."""
    username = input("Enter a username: ").lower()
    sheet, wb = load_sheet("Users")
    if not sheet:
        print("Error loading user data.")
        return

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0].lower() == username:
            print("Username already exists. Try a different one.")
            return

    password = getpass("Enter a strong password: ")
    if not is_strong_password(password):
        print("Weak password. Try again.")
        return

    confirm_password = getpass("Confirm your password: ")
    if password != confirm_password:
        print("Passwords do not match. Try again.")
        return

    role = "admin" if input("Is this user an admin? (yes/no): ").strip().lower() == "yes" else "user"
    if role == "admin":
        admin_password = getpass("Enter the admin confirmation password: ")
        if hash_password(admin_password) != ADMIN_CONFIRM_PASSWORD:
            print("Invalid admin confirmation password. Registration failed.")
            return

    hashed_password = hash_password(password)
    sheet.append([username, hashed_password, role])
    save_to_excel(wb)
    print(f"{role.capitalize()} '{username}' registered successfully.")

def login_user():
    """Log in a user and return their username and role."""
    username = input("Enter your username: ").strip().lower()
    password = getpass("Enter your password: ")
    hashed_password = hash_password(password)

    sheet, _ = load_sheet("Users")
    if not sheet:
        print("Error loading user data.")
        return None, None

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0].lower() == username and row[1] == hashed_password:
            otp = generate_otp()
            print(f"Your OTP is: {otp}")
            user_otp = int(input("Enter the OTP: "))
            if verify_otp(user_otp, otp):
                print(f"Welcome, {username}! You are logged in as {row[2]}.")
                return username, row[2]
            else:
                print("Invalid OTP. Login failed.")
                return None, None

    print("Invalid username or password.")
    return None, None

def generate_report_by_date():
    """Generate attendance report filtered by date."""
    start_date = input("Enter the start date (YYYY-MM-DD): ").strip()
    end_date = input("Enter the end date (YYYY-MM-DD): ").strip()

    sheet, _ = load_sheet("Attendance")
    if not sheet:
        print("Error loading attendance data.")
        return

    print(f"Attendance from {start_date} to {end_date}:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Assume date is stored in the third column
        attendance_date = row[2][:10]  # Extract date portion
        if start_date <= attendance_date <= end_date:
            print(row)


def add_student_attendance(username):
    """Allow students to add their attendance."""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet, wb = load_sheet("Attendance")
    if not sheet:
        print("Error loading attendance data.")
        return

    # Check if student already checked in
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username and not row[3]:  # No check-out time yet
            print("You have already checked in.")
            return

    sheet.append([username, datetime.now().date(), timestamp, "", ""])
    save_to_excel(wb)
    print(f"Attendance marked for {username} at {timestamp}.")

def view_attendance_from_excel():
    """View attendance records from the Excel sheet."""
    sheet, _ = load_sheet("Attendance")
    if not sheet:
        print("Error loading attendance data.")
        return

    print("\nAttendance Records:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Username: {row[0]}, Attendance: {row[1]}")    

def view_attendance_list():
    """Display the attendance list."""
    sheet, _ = load_sheet("Attendance")
    if not sheet:
        print("Error loading attendance data.")
        return

    print("\n--- Attendance List ---")
    print(f"{'Username':<20}{'Date':<15}{'Check-in-Time':<20}{'Check-out-Time':<20}{'Total Hours':<15}")
    print("-" * 90)

    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Ensure row is not empty
                username = row[0] or "N/A"
                date = row[1] or "N/A"
                check_in_time = row[2] or "N/A"
                check_out_time = row[3] or "N/A"
                total_hours = row[4] or "N/A"

                print(f"{username:<20}{date:<15}{check_in_time:<20}{check_out_time:<20}{total_hours:<15}")
    except Exception as e:
        logging.error(f"Error displaying attendance list: {e}")
        print("An unexpected error occurred while displaying attendance. Please check the logs.")

@require_role("admin")
def admin_menu(username, role):
    """Menu for admins to manage attendance."""
    while True:
        print("\n--- Admin Menu ---")
        print("1. Add Attendance Record")
        print("2. Remove Attendance Record")
        print("3. View Attendance List")
        print("4. Generate Attendance Report")
        print("5. Back to Main Menu")
        choice = input("Choose an option: ").strip()

        if choice == "1":
            add_student_attendance(username)
        elif choice == "2":
            remove_from_attendance_list()
        elif choice == "3":
            view_attendance_list()
        elif choice == "4":
            generate_report_by_date()
        elif choice == "5":
            break
        else:
            print("Invalid option. Try again.")

def log_audit_action(username, action, details=""):
    """Log administrative actions with timestamps."""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet, wb = load_sheet("Audit Log")

    # Initialize Audit Log sheet if it doesn't exist
    if not sheet:
        sheet, wb = load_sheet("Audit Log")
        if not sheet:
            print("Error loading audit log.")
            return

        sheet.append(["Timestamp", "Username", "Action", "Details"])

    sheet.append([timestamp, username, action, details])
    save_to_excel(wb)
    logging.info(f"{username} performed action: {action} at {timestamp}. Details: {details}")


def add_to_attendance_list(username):
    """Add a student to the attendance list."""
    student_name = input("Enter the student's name to add: ").strip()
    attendance_list.append(student_name)
    print(f"{student_name} has been added to the attendance list.")
    update_attendance_in_excel(username)

def remove_from_attendance_list():
    """Remove a student from the attendance list."""
    student_name = input("Enter the student's name to remove: ").strip()
    if student_name in attendance_list:
        attendance_list.remove(student_name)
        print(f"{student_name} has been removed from the attendance list.")
    else:
        print(f"{student_name} is not in the attendance list.")

def update_attendance_in_excel(username):
    """Update attendance data in the Excel sheet."""
    sheet, wb = load_sheet("Attendance")
    if not sheet:
        print("Error loading attendance data.")
        return

    sheet.append([username, ", ".join(attendance_list)])
    save_to_excel(wb)
    print("Attendance has been updated in the Excel sheet.")


def reset_password_via_email():
    """Reset user password and send a temporary password via email."""
    username = input("Enter the username for password reset: ").strip().lower()
    sheet, wb = load_sheet("Users")
    if not sheet:
        print("Error accessing user data.")
        return

    user_email = input("Enter the user's registered email: ").strip()
    temp_password = f"Temp{random.randint(1000, 9999)}"
    hashed_temp_password = hash_password(temp_password)

    for row in sheet.iter_rows(min_row=2):
        if row[0].value.lower() == username:
            row[1].value = hashed_temp_password
            save_to_excel(wb)
            subject = "Password Reset Notification"
            message = f"Your temporary password is: {temp_password}\nPlease change it after logging in."
            try:
                send_email_notification(user_email, subject, message)
                print(f"Temporary password sent to {user_email}")
            except Exception as e:
                print(f"Failed to send email: {e}")
            return

    print("Username not found.")

def main_menu():
    """Main menu for the application."""
    try:
        initialize_excel()
        while True:
            print("\nMain Menu")
            print("1. Register User")
            print("2. Login")
            print("3. Exit")
            choice = input("Choose an option: ").strip()

            if choice == "1":
                register_user()
            elif choice == "2":
                username, role = login_user()
                if username and role:
                    if role == "admin":
                        manage_attendance(username, role)
                    else:
                        print("Attendance tracking for regular users is under development.")
            elif choice == "3":
                if input("Are you sure you want to exit? (yes/no): ").strip().lower() == "yes":
                    print("Goodbye!")
                    break
            else:
                print("Invalid option. Try again.")
    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        print("An unexpected error occurred. Please check the logs.")

def main_menu():
    """Main menu for the application."""
    try:
        initialize_excel()
        while True:
            print("\n--- Main Menu ---")
            print("1. Register User")
            print("2. Login")
            print("3. Exit")
            choice = input("Choose an option: ").strip()

            if choice == "1":
                register_user()
            elif choice == "2":
                username, role = login_user()
                if username and role:
                    if role == "admin":
                        admin_menu(username, role)
                    elif role == "user":
                        student_menu(username)
            elif choice == "3":
                if input("Are you sure you want to exit? (yes/no): ").strip().lower() == "yes":
                    print("Goodbye!")
                    break
            else:
                print("Invalid option. Try again.")
    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        print("An unexpected error occurred. Please check the logs.")

def student_menu(username):
    """Menu for students to manage their attendance."""
    while True:
        print("\n--- Student Menu ---")
        print("1. Add My Attendance")
        print("2. View Attendance List")
        print("3. Back to Main Menu")
        choice = input("Choose an option: ").strip()

        if choice == "1":
            add_to_attendance_list(username)
        elif choice == "2":
            view_attendance_list()
        elif choice == "3":
            break
        else:
            print("Invalid option. Try again.")

@require_role("admin")
def admin_menu(username, role):
    """Menu for admins to manage attendance."""
    while True:
        print("\n--- Admin Menu ---")
        print("1. Add Student to Attendance")
        print("2. View Attendance List")
        print("3. Remove Student from Attendance")
        print("4. Back to Main Menu")
        choice = input("Choose an option: ").strip()

        if choice == "1":
            add_to_attendance_list(username)
        elif choice == "2":
            view_attendance_list()
        elif choice == "3":
            remove_from_attendance_list()
        elif choice == "4":
            break
        else:
            print("Invalid option. Try again.")


if __name__ == "__main__":
    main_menu()
